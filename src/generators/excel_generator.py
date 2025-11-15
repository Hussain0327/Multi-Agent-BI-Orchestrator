"""
Excel workbook generation from structured agent outputs.

Creates professional analysis workbooks with multiple sheets:
- Executive Summary (KPI dashboard)
- Raw Data
- Calculations (formulas, scenarios)
- Charts & Visuals
- Assumptions & Sources
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from pathlib import Path
from typing import Optional
from datetime import datetime

from src.schemas import AgentOutput, Metric, Recommendation


class ExcelGenerator:
    """Generate Excel workbooks from agent outputs."""

    def __init__(self):
        """Initialize Excel generator."""
        # Color scheme (matching Valtric theme)
        self.colors = {
            'primary': '2C3E50',
            'secondary': '3498DB',
            'success': '2ECC71',
            'warning': 'F39C12',
            'accent': 'E74C3C',
            'light_gray': 'ECF0F1',
            'dark_gray': '7F8C8D',
        }

    def generate(
        self,
        agent_output: AgentOutput,
        output_path: Optional[str] = None,
        include_scenarios: bool = True
    ) -> str:
        """
        Generate Excel workbook from agent output.

        Args:
            agent_output: Structured output from agent
            output_path: Where to save .xlsx file
            include_scenarios: Include scenario analysis (Base/Upside/Downside)

        Returns:
            Path to saved Excel file
        """
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Add sheets
        self._add_executive_summary(wb, agent_output)
        self._add_raw_data_sheet(wb, agent_output)
        self._add_calculations_sheet(wb, agent_output, include_scenarios)
        self._add_charts_sheet(wb, agent_output)
        self._add_assumptions_sheet(wb, agent_output)

        # Save
        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"analysis_{agent_output.agent}_{timestamp}.xlsx"

        wb.save(output_path)
        print(f"✓ Excel workbook saved: {output_path}")
        return output_path

    def _add_executive_summary(self, wb: Workbook, output: AgentOutput):
        """Add executive summary sheet with KPI dashboard."""
        ws = wb.create_sheet("Executive Summary", 0)

        # Title
        ws['A1'] = "Executive Summary"
        ws['A1'].font = Font(size=18, bold=True, color=self.colors['primary'])
        ws.merge_cells('A1:D1')

        # Query
        ws['A3'] = "Analysis Question:"
        ws['A3'].font = Font(bold=True)
        ws['B3'] = output.query
        ws.merge_cells('B3:D3')

        # Executive summary text
        ws['A5'] = "Summary:"
        ws['A5'].font = Font(bold=True)
        ws['A6'] = output.findings.executive_summary
        ws.merge_cells('A6:D10')
        ws['A6'].alignment = Alignment(wrap_text=True, vertical='top')

        # Key metrics section
        row = 12
        ws[f'A{row}'] = "Key Metrics"
        ws[f'A{row}'].font = Font(size=14, bold=True, color=self.colors['primary'])
        row += 1

        # Headers
        headers = ['Metric', 'Value', 'Unit', 'Confidence']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.colors['secondary'],
                                   end_color=self.colors['secondary'],
                                   fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        row += 1

        # Metrics data
        for name, metric in output.findings.metrics.items():
            ws.cell(row, 1, name.replace('_', ' ').title())
            ws.cell(row, 2, metric.value)
            ws.cell(row, 3, metric.unit)
            ws.cell(row, 4, metric.confidence.title())

            # Color code confidence
            conf_cell = ws.cell(row, 4)
            if metric.confidence == 'high':
                conf_cell.fill = PatternFill(start_color=self.colors['success'],
                                            end_color=self.colors['success'],
                                            fill_type='solid')
            elif metric.confidence == 'medium':
                conf_cell.fill = PatternFill(start_color=self.colors['warning'],
                                            end_color=self.colors['warning'],
                                            fill_type='solid')

            row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15

    def _add_raw_data_sheet(self, wb: Workbook, output: AgentOutput):
        """Add raw data sheet with all structured outputs."""
        ws = wb.create_sheet("Raw Data")

        # Title
        ws['A1'] = "Raw Analysis Data"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:B1')

        row = 3

        # Metadata
        ws[f'A{row}'] = "Metadata"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        metadata_items = [
            ('Query', output.query),
            ('Agent', output.agent.replace('_', ' ').title()),
            ('Timestamp', output.timestamp.strftime('%Y-%m-%d %H:%M:%S')),
            ('Model', output.metadata.model),
            ('Confidence', output.metadata.confidence.title()),
        ]

        for label, value in metadata_items:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = str(value)
            row += 1

        row += 2

        # Key Findings
        ws[f'A{row}'] = "Key Findings"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        for i, finding in enumerate(output.findings.key_findings, 1):
            ws[f'A{row}'] = f"{i}."
            ws[f'B{row}'] = finding
            ws.merge_cells(f'B{row}:D{row}')
            ws[f'B{row}'].alignment = Alignment(wrap_text=True)
            row += 1

        row += 2

        # Risks
        if output.findings.risks:
            ws[f'A{row}'] = "Risks & Considerations"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1

            for i, risk in enumerate(output.findings.risks, 1):
                ws[f'A{row}'] = f"{i}."
                ws[f'B{row}'] = risk
                ws.merge_cells(f'B{row}:D{row}')
                ws[f'B{row}'].alignment = Alignment(wrap_text=True)
                row += 1

        row += 2

        # Recommendations
        if output.findings.recommendations:
            ws[f'A{row}'] = "Recommendations"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1

            for rec in output.findings.recommendations:
                ws[f'A{row}'] = rec.title
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'] = f"[{rec.priority.upper()}]"

                # Color code priority
                priority_cell = ws[f'B{row}']
                if rec.priority == 'high':
                    priority_cell.fill = PatternFill(start_color=self.colors['accent'],
                                                     end_color=self.colors['accent'],
                                                     fill_type='solid')
                    priority_cell.font = Font(color='FFFFFF', bold=True)

                row += 1
                ws[f'A{row}'] = "Impact:"
                ws[f'B{row}'] = rec.impact
                ws.merge_cells(f'B{row}:D{row}')
                row += 1
                ws[f'A{row}'] = "Actions:"
                row += 1

                for action in rec.action_items:
                    ws[f'B{row}'] = f"• {action}"
                    ws.merge_cells(f'B{row}:D{row}')
                    ws[f'B{row}'].alignment = Alignment(wrap_text=True)
                    row += 1

                row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20

    def _add_calculations_sheet(self, wb: Workbook, output: AgentOutput, include_scenarios: bool):
        """Add calculations sheet with formulas and scenario analysis."""
        ws = wb.create_sheet("Calculations")

        # Title
        ws['A1'] = "Financial Calculations"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:E1')

        row = 3

        # Metrics with formulas
        ws[f'A{row}'] = "Metric"
        ws[f'B{row}'] = "Base Case"
        ws[f'C{row}'] = "Upside (+20%)"
        ws[f'D{row}'] = "Downside (-20%)"
        ws[f'E{row}'] = "Formula"

        for col in range(1, 6):
            cell = ws.cell(row, col)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.colors['primary'],
                                   end_color=self.colors['primary'],
                                   fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        row += 1
        base_row = row

        # Add metrics with scenario formulas
        for name, metric in output.findings.metrics.items():
            ws.cell(row, 1, name.replace('_', ' ').title())

            # Base case
            base_cell = ws.cell(row, 2, metric.value)

            if isinstance(metric.value, (int, float)):
                # Upside scenario (+20%)
                ws.cell(row, 3, f"=B{row}*1.2")
                # Downside scenario (-20%)
                ws.cell(row, 4, f"=B{row}*0.8")
                # Formula reference
                ws.cell(row, 5, metric.formula if metric.formula else "Direct input")
            else:
                # Non-numeric values
                ws.cell(row, 3, metric.value)
                ws.cell(row, 4, metric.value)
                ws.cell(row, 5, "N/A")

            row += 1

        # Add some calculated metrics if we have key financial data
        if 'LTV' in output.findings.metrics and 'CAC' in output.findings.metrics:
            row += 1
            ws[f'A{row}'] = "Key Ratios"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1

            # Find LTV and CAC rows
            ltv_row = None
            cac_row = None
            for r in range(base_row, row):
                cell_value = ws.cell(r, 1).value
                if cell_value and 'LTV' in str(cell_value).upper():
                    ltv_row = r
                if cell_value and 'CAC' in str(cell_value).upper():
                    cac_row = r

            if ltv_row and cac_row:
                ws[f'A{row}'] = "LTV:CAC Ratio"
                ws[f'B{row}'] = f"=B{ltv_row}/B{cac_row}"
                ws[f'C{row}'] = f"=C{ltv_row}/C{cac_row}"
                ws[f'D{row}'] = f"=D{ltv_row}/D{cac_row}"
                ws[f'E{row}'] = "=LTV/CAC"

                # Format as ratio
                for col in [2, 3, 4]:
                    ws.cell(row, col).number_format = '0.00'

        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 30

    def _add_charts_sheet(self, wb: Workbook, output: AgentOutput):
        """Add charts and visualizations sheet."""
        ws = wb.create_sheet("Charts & Visuals")

        # Title
        ws['A1'] = "Data Visualizations"
        ws['A1'].font = Font(size=16, bold=True)

        # Note: In a full implementation, we would embed matplotlib charts here
        # For now, we'll add chart data and simple Excel charts

        ws['A3'] = "Chart data will be generated from metrics"
        ws['A4'] = "In production, matplotlib charts can be embedded as images"

    def _add_assumptions_sheet(self, wb: Workbook, output: AgentOutput):
        """Add assumptions and sources sheet."""
        ws = wb.create_sheet("Assumptions & Sources")

        # Title
        ws['A1'] = "Assumptions & Data Sources"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:C1')

        row = 3

        # Assumptions section
        ws[f'A{row}'] = "Assumptions"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        ws[f'A{row}'] = "Metric"
        ws[f'B{row}'] = "Source"
        ws[f'C{row}'] = "Confidence"

        for col in [1, 2, 3]:
            cell = ws.cell(row, col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=self.colors['light_gray'],
                                   end_color=self.colors['light_gray'],
                                   fill_type='solid')

        row += 1

        # List assumptions from metrics
        for name, metric in output.findings.metrics.items():
            ws.cell(row, 1, name.replace('_', ' ').title())
            ws.cell(row, 2, metric.source.replace('_', ' ').title())
            ws.cell(row, 3, metric.confidence.title())
            row += 1

        row += 2

        # Citations section
        if output.research_citations:
            ws[f'A{row}'] = "Research Citations"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1

            for citation in output.research_citations:
                authors_str = ', '.join(citation.authors[:3])
                if len(citation.authors) > 3:
                    authors_str += ' et al.'

                citation_text = f"{authors_str} ({citation.year}). {citation.title}."
                ws[f'A{row}'] = citation_text
                ws.merge_cells(f'A{row}:C{row}')
                ws[f'A{row}'].alignment = Alignment(wrap_text=True)
                row += 1

                if citation.url:
                    ws[f'A{row}'] = f"URL: {citation.url}"
                    ws[f'A{row}'].font = Font(color=self.colors['secondary'], underline='single')
                    row += 1

                row += 1

        row += 2

        # Methodology section
        ws[f'A{row}'] = "Analysis Methodology"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        methodology_items = [
            f"Agent: {output.agent.replace('_', ' ').title()}",
            f"Model: {output.metadata.model}",
            f"Analysis Date: {output.timestamp.strftime('%Y-%m-%d %H:%M:%S')}",
            f"Overall Confidence: {output.metadata.confidence.title()}",
        ]

        for item in methodology_items:
            ws[f'A{row}'] = item
            row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15


    def generate_from_json(self, json_path: str, output_path: Optional[str] = None) -> str:
        """
        Generate Excel workbook from JSON file.

        Args:
            json_path: Path to agent output JSON file
            output_path: Where to save Excel file

        Returns:
            Path to saved Excel file
        """
        import json
        with open(json_path, 'r') as f:
            data = json.load(f)

        from src.schemas import AgentOutput
        agent_output = AgentOutput(**data)

        return self.generate(agent_output, output_path)
